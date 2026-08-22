"""Reports: monthly sheet, employee detail, late report, department summary and exports."""
import csv
from calendar import monthrange
from datetime import date as date_cls, datetime, timedelta
from io import BytesIO

from django.db.models import Avg, Count, Q, Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from apps.utils import now as app_now, today as app_today

from apps.accounts.models import Department, User
from apps.accounts.permissions import manager_required
from apps.attendance.models import DailyAttendance, Outing
from apps.attendance.services import month_summary
from apps.tasks.models import DailyTask

S = DailyAttendance.Status


def _month_range(request):
    today = app_today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))
    first = date_cls(year, month, 1)
    last = date_cls(year, month, monthrange(year, month)[1])
    return year, month, first, last


@manager_required
def report_home(request):
    today = app_today()
    return render(request, "reports/report_home.html", {
        "today": today,
        "month_label": today.strftime("%B %Y"),
    })


@manager_required
def monthly_sheet(request):
    """Employee x day grid - the whole month at a glance."""
    year, month, first, last = _month_range(request)
    dept = request.GET.get("dept", "")

    employees = User.objects.active_staff().select_related("department")
    if dept:
        employees = employees.filter(department_id=dept)

    records = DailyAttendance.objects.filter(date__range=(first, last), employee__in=employees)
    lookup = {(r.employee_id, r.date): r for r in records}
    days = [first + timedelta(days=i) for i in range((last - first).days + 1)]

    rows = []
    for emp in employees:
        cells, totals = [], {"present": 0, "late": 0, "absent": 0, "leave": 0, "home": 0, "minutes": 0}
        for d in days:
            r = lookup.get((emp.id, d))
            cells.append({"date": d, "record": r})
            if not r:
                continue
            totals["minutes"] += r.worked_minutes
            if r.is_late:
                totals["late"] += 1
            if r.status == S.ABSENT:
                totals["absent"] += 1
            elif r.status == S.LEAVE:
                totals["leave"] += 1
            elif r.status not in (S.WEEKEND, S.HOLIDAY):
                totals["present"] += 1
            if r.mode == DailyAttendance.Mode.HOME:
                totals["home"] += 1
        totals["hours"] = round(totals["minutes"] / 60, 1)
        rows.append({"employee": emp, "cells": cells, "totals": totals})

    return render(request, "reports/monthly_sheet.html", {
        "rows": rows, "days": days, "year": year, "month": month,
        "month_label": first.strftime("%B %Y"), "dept": dept,
        "departments": Department.objects.all(),
        "prev": first - timedelta(days=1), "next": last + timedelta(days=1),
    })


@manager_required
def employee_report(request):
    """One employee's detailed monthly report."""
    year, month, first, last = _month_range(request)
    emp_id = request.GET.get("employee")
    employees = User.objects.active_staff()
    employee = employees.filter(pk=emp_id).first() or employees.first()
    records, summary, tasks = [], None, []
    if employee:
        records = DailyAttendance.objects.filter(
            employee=employee, date__range=(first, last)).order_by("date")
        summary = month_summary(employee, year, month)
        tasks = DailyTask.objects.filter(employee=employee, date__range=(first, last))
    return render(request, "reports/employee_report.html", {
        "employees": employees, "employee": employee, "records": records,
        "summary": summary, "tasks": tasks, "year": year, "month": month,
        "month_label": first.strftime("%B %Y"),
    })


@manager_required
def late_report(request):
    year, month, first, last = _month_range(request)
    rows = (DailyAttendance.objects.filter(date__range=(first, last))
            .values("employee__id", "employee__employee_id", "employee__first_name",
                    "employee__last_name", "employee__department__name")
            .annotate(
                late_days=Count("id", filter=Q(is_late=True)),
                early_days=Count("id", filter=Q(is_early_out=True)),
                absent_days=Count("id", filter=Q(status=S.ABSENT)),
                missing_out=Count("id", filter=Q(status=S.MISSING_OUT)),
                total_late_min=Sum("late_minutes"),
                total_short_min=Sum("shortfall_minutes"),
                avg_worked=Avg("worked_minutes"),
            ).order_by("-late_days", "-total_late_min"))
    return render(request, "reports/late_report.html", {
        "rows": rows, "month_label": first.strftime("%B %Y"),
        "year": year, "month": month,
    })


@manager_required
def department_report(request):
    year, month, first, last = _month_range(request)
    rows = []
    for dept in Department.objects.all():
        qs = DailyAttendance.objects.filter(date__range=(first, last), employee__department=dept)
        total = qs.exclude(status__in=[S.WEEKEND, S.HOLIDAY]).count() or 1
        rows.append({
            "dept": dept,
            "headcount": User.objects.active_staff().filter(department=dept).count(),
            "on_time": qs.filter(status=S.PRESENT).count(),
            "late": qs.filter(is_late=True).count(),
            "absent": qs.filter(status=S.ABSENT).count(),
            "home": qs.filter(mode=DailyAttendance.Mode.HOME).count(),
            "hours": round((qs.aggregate(m=Sum("worked_minutes"))["m"] or 0) / 60, 1),
            "punctuality": round(qs.filter(status=S.PRESENT).count() / total * 100),
        })
    return render(request, "reports/department_report.html", {
        "rows": rows, "month_label": first.strftime("%B %Y"), "year": year, "month": month,
    })


# ------------------------------------------------------------------ exports
def _export_rows(first, last):
    qs = (DailyAttendance.objects.filter(date__range=(first, last))
          .select_related("employee", "employee__department").order_by("date", "employee__employee_id"))
    yield ["Date", "Employee ID", "Name", "Department", "Mode", "In", "Required out",
           "Out", "Worked", "Late (min)", "Early out (min)", "Overtime (min)", "Status", "Remarks"]
    for r in qs:
        yield [
            r.date.strftime("%Y-%m-%d"), r.employee.employee_id, r.employee.display_name,
            r.employee.department.name if r.employee.department else "",
            r.get_mode_display(),
            r.check_in.strftime("%I:%M %p") if r.check_in else "",
            r.required_out.strftime("%I:%M %p") if r.required_out else "",
            r.check_out.strftime("%I:%M %p") if r.check_out else "",
            r.worked_display, r.late_minutes, r.early_out_minutes, r.overtime_minutes,
            r.get_status_display(), r.remarks,
        ]


@manager_required
def export_csv(request):
    year, month, first, last = _month_range(request)
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="attendance-{year}-{month:02d}.csv"'
    writer = csv.writer(response)
    for row in _export_rows(first, last):
        writer.writerow(row)
    return response


@manager_required
def export_excel(request):
    year, month, first, last = _month_range(request)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return HttpResponse("openpyxl is not installed. Run: pip install openpyxl", status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{first:%b %Y}"
    header_fill = PatternFill("solid", fgColor="14343B")
    for i, row in enumerate(_export_rows(first, last)):
        ws.append(row)
        if i == 0:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
    widths = [12, 12, 22, 16, 12, 11, 13, 11, 10, 10, 12, 12, 14, 40]
    for col, w in zip(ws.columns, widths):
        ws.column_dimensions[col[0].column_letter].width = w
    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="attendance-{year}-{month:02d}.xlsx"'
    return response


@manager_required
def outing_report(request):
    """Where the team went during working hours, and for how long."""
    today = app_today()
    day_from = request.GET.get("from")
    day_to = request.GET.get("to")
    try:
        day_from = datetime.strptime(day_from, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        day_from = today - timedelta(days=13)
    try:
        day_to = datetime.strptime(day_to, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        day_to = today

    purpose = request.GET.get("purpose", "")
    emp_id = request.GET.get("employee", "")
    dept = request.GET.get("dept", "")

    outings = (Outing.objects
               .filter(date__range=(day_from, day_to))
               .select_related("employee", "employee__department"))
    if purpose:
        outings = outings.filter(purpose=purpose)
    if emp_id:
        outings = outings.filter(employee_id=emp_id)
    if dept:
        outings = outings.filter(employee__department_id=dept)
    outings = outings.order_by("-date", "left_at")

    per_employee = (outings.values(
                        "employee__id", "employee__employee_id",
                        "employee__first_name", "employee__last_name")
                    .annotate(trips=Count("id"),
                              total_minutes=Sum("minutes"),
                              official=Count("id", filter=Q(purpose=Outing.Purpose.OFFICIAL)),
                              personal=Count("id", filter=Q(purpose=Outing.Purpose.PERSONAL)),
                              unstated=Count("id", filter=Q(purpose=Outing.Purpose.UNSPECIFIED)))
                    .order_by("-total_minutes"))

    totals = outings.aggregate(
        trips=Count("id"),
        total_minutes=Sum("minutes"),
        official_minutes=Sum("minutes", filter=Q(purpose=Outing.Purpose.OFFICIAL)),
        personal_minutes=Sum("minutes", filter=Q(purpose=Outing.Purpose.PERSONAL)),
        unstated=Count("id", filter=Q(purpose=Outing.Purpose.UNSPECIFIED)),
    )

    return render(request, "reports/outing_report.html", {
        "outings": outings[:400],
        "per_employee": per_employee,
        "totals": totals,
        "day_from": day_from, "day_to": day_to,
        "purpose": purpose, "emp_id": emp_id, "dept": dept,
        "purposes": Outing.Purpose.choices,
        "employees": User.objects.active_staff(),
        "departments": Department.objects.all(),
        "out_now": outings.filter(returned_at__isnull=True).count(),
    })


@manager_required
def export_outings_csv(request):
    """The same trips as a spreadsheet."""
    today = app_today()
    try:
        day_from = datetime.strptime(request.GET.get("from"), "%Y-%m-%d").date()
    except (TypeError, ValueError):
        day_from = today - timedelta(days=13)
    try:
        day_to = datetime.strptime(request.GET.get("to"), "%Y-%m-%d").date()
    except (TypeError, ValueError):
        day_to = today

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="trips-{day_from}-to-{day_to}.csv"'
    writer = csv.writer(response)
    writer.writerow(["Date", "Employee ID", "Name", "Department", "Left at", "Returned at",
                     "Minutes", "Purpose", "Counts as work", "Destination", "Note"])
    for o in (Outing.objects.filter(date__range=(day_from, day_to))
              .select_related("employee", "employee__department").order_by("date", "left_at")):
        writer.writerow([
            o.date.strftime("%Y-%m-%d"), o.employee.employee_id, o.employee.display_name,
            o.employee.department.name if o.employee.department else "",
            o.left_at.strftime("%I:%M %p"),
            o.returned_at.strftime("%I:%M %p") if o.returned_at else "still out",
            o.minutes, o.get_purpose_display(), "Yes" if o.counts_as_work else "No",
            o.destination, o.note,
        ])
    return response
